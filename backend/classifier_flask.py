import pickle

import openpyxl
from flask import Flask, jsonify, make_response, request
import requests
from bs4 import BeautifulSoup
import pandas as pd
from flask_cors import cross_origin, CORS

app = Flask(__name__)
CORS(app, support_credentials=True)
app.config['JSON_AS_ASCII'] = False

vectorizer = pickle.load(open('models/vectorizer.sav', 'rb'))
classifier = pickle.load(open('models/classifier.sav', 'rb'))

header = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/77.0.3865.90 Safari/537.36'}
cookie = {}


def search_asin(asin):
    url = "https://www.amazon.in/dp/" + asin
    page = requests.get(url, cookies=cookie, headers=header)
    if page.status_code == 200:
        return page
    else:
        return "Error"


def search_reviews(review_link):
    url = "https://www.amazon.in" + review_link
    page = requests.get(url, cookies=cookie, headers=header)
    if page.status_code == 200:
        return page
    else:
        return "Error"


@app.route('/sentimentbylink', methods=['GET', 'POST'])
def sentiment_analysis_of_reviews():
    if request.method == 'GET':
        link = request.args.get('link')
        asin = link.split("/dp/")[1]
        asin = asin[0:10]

        link = []
        data = []
        data.append(asin)

        for i in range(len(data)):
            response = search_asin(data[i])
            if isinstance(response, str):
                print('Error in fetching data')
                break
            soup = BeautifulSoup(response.content, features="html.parser")
            for j in soup.findAll("a", {'data-hook': "see-all-reviews-link-foot"}):
                link.append(j['href'])

        reviews = []
        url_link = link[0]
        for k in range(100):
            response = search_reviews(url_link + '&pageNumber=' + str(k))
            if isinstance(response, str):
                print('Error in fetching data')
                break
            soup = BeautifulSoup(response.content, features="html.parser")
            for i in soup.findAll("span", {'data-hook': "review-body"}):
                reviews.append(i.text)

        rev = {'reviews': reviews}
        review_data = pd.DataFrame.from_dict(rev)
        review_data.to_excel('Scraping reviews.xlsx', index=False)

        excel_file = openpyxl.load_workbook("Scraping reviews.xlsx")
        file_iterator = excel_file.active

        positive_reviews_array = []
        negative_reviews_array = []

        for rows in range(2, file_iterator.max_row + 1):
            review_row_wise = ""
            for columns in range(1, file_iterator.max_column + 1):
                review_text = file_iterator.cell(row=rows, column=columns)
                if isinstance(review_text.value, str):
                    review_row_wise += review_text.value
            if review_row_wise:
                text_vector = vectorizer.transform([review_row_wise])
                result = classifier.predict(text_vector)
                if result[0] == "pos":
                    positive_reviews_array.append(review_row_wise)
                else:
                    negative_reviews_array.append(review_row_wise)

        return jsonify({
            "status": 200,
            "positive_reviews": positive_reviews_array,
            "negative_reviews": negative_reviews_array
        })


@cross_origin(supports_credentials=True)
@app.route('/sentimentbyonereview', methods=['GET', 'POST'])
def sentiment_analysis():
    if request.method == 'GET':
        text = request.args.get('text')
        if text:
            text_vector = vectorizer.transform([text])
            result = classifier.predict(text_vector)
            if result[0] == "pos":
                result[0] = "positive"
            else:
                result[0] = "negative"
            return jsonify({
                "status": 200,
                "classificationType": result[0],
                "reviewInput": text
            })
        return make_response(jsonify({'error': 'sorry! unable to parse', 'status_code': 500}), 500)


if __name__ == '__main__':
    app.run()
